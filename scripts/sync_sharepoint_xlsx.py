"""
sync_sharepoint_xlsx.py
Pipeline automático: SharePoint → JSON para o Portal BMax.

Fluxo:
  1. Autentica via MSAL (token cache do CI ou local)
  2. Lista todos os xlsx da pasta SharePoint do BMax
  3. Baixa cada arquivo
  4. Converte para JSON (estrutura específica por planilha)
  5. Salva em src/data/*.json

Env vars:
  MSAL_TOKEN_CACHE  — cache JSON (GitHub Secret ou .oauth_token_cache.json local)

Uso:
  python scripts/sync_sharepoint_xlsx.py
"""

import os, sys, json, io
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent.parent / ".env")
except ImportError:
    pass

TENANT_ID = "c0bbec8e-9949-4dcf-80b3-bd21689c33e4"
CLIENT_ID = "d1fb2af8-a56b-41a0-8401-e87591360016"
SCOPE     = ["https://graph.microsoft.com/Sites.Read.All"]
DRIVE_ID  = "b!6Kvw0SoDTUiClr8MvMu7NuwGBC9zdONMq_ED-vRKE3W43Ls7sVoDTZSN_i_i5kpD"

SP_FOLDER = "COMERCIAL/Gestor/Salescope/Projeto BMax/boxer-portal-bmax/src/data"

DATA_DIR = Path(__file__).parent.parent / "src" / "data"


def get_graph_token() -> str:
    import msal
    cache_env = os.environ.get("MSAL_TOKEN_CACHE", "")
    cache = msal.SerializableTokenCache()
    if cache_env:
        cache.deserialize(cache_env)
    else:
        local = Path(__file__).parent.parent.parent.parent.parent.parent / \
                "App Vendas/bsis_v3/backend/data/.oauth_token_cache.json"
        if local.exists():
            cache.deserialize(local.read_text(encoding="utf-8"))
    app = msal.PublicClientApplication(
        CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{TENANT_ID}",
        token_cache=cache,
    )
    accounts = app.get_accounts()
    if not accounts:
        raise RuntimeError("MSAL: nenhuma conta em cache. Renove com renovar_msal_ci.py")
    result = app.acquire_token_silent(SCOPE, account=accounts[0])
    if not result or "access_token" not in result:
        raise RuntimeError(f"MSAL: token expirado — {result.get('error_description','')}")
    return result["access_token"]


def list_xlsx_files(token: str) -> list:
    import requests
    url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/root:/{SP_FOLDER}:/children"
    hdrs = {"Authorization": f"Bearer {token}"}
    r = requests.get(url, headers=hdrs, timeout=30)
    if r.status_code == 404:
        print(f"WARN: Pasta SharePoint não encontrada: {SP_FOLDER}")
        print("Tentando download direto por item ID...")
        return None
    r.raise_for_status()
    items = r.json().get("value", [])
    return [i for i in items if i["name"].lower().endswith(".xlsx")]


def download_file_by_id(token: str, item_id: str) -> bytes:
    import requests
    url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{item_id}/content"
    r = requests.get(url, headers={"Authorization": f"Bearer {token}"}, allow_redirects=True, timeout=60)
    r.raise_for_status()
    return r.content


def download_file_by_path(token: str, path: str) -> bytes:
    import requests
    url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/root:/{path}:/content"
    r = requests.get(url, headers={"Authorization": f"Bearer {token}"}, allow_redirects=True, timeout=60)
    r.raise_for_status()
    return r.content


KNOWN_FILES = {
    "BMAX CRITERIOS V2.xlsx": "015WUCIWDTZYXSMZAJGJGIUOY4XZKPVHD6",
}


def convert_ibge_to_json(xlsx_bytes: bytes) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb["Base IBGE"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    entries = []
    for row in rows:
        if len(row) < 21:
            continue
        municipio = row[16]
        estado = row[1]
        responsavel = row[20]
        if municipio and estado and responsavel and str(municipio).strip() != "Nome_Distrito":
            entries.append({
                "municipio": str(municipio).strip(),
                "estado": str(estado).strip(),
                "responsavel": str(responsavel).strip()
            })
    wb.close()
    return {"data": entries, "count": len(entries)}


def convert_criterios_to_json(xlsx_bytes: bytes) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb["PCI Versão 2 (atual)"]

    header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    pcis = {}
    for col_idx, val in enumerate(header_row):
        if col_idx >= 6 and val:
            pci_name = str(val).strip().upper().replace(" ", "")
            pcis[pci_name] = col_idx

    cashback = {"revenda": {}, "representante": {}}

    all_rows = list(ws.iter_rows(values_only=True))

    for pci_name, col_idx in pcis.items():
        rev_val = all_rows[17][col_idx] if len(all_rows) > 17 else None
        rep_val = all_rows[15][col_idx] if len(all_rows) > 15 else None

        if isinstance(rev_val, str) and "-" in rev_val:
            rev_by_classe = {}
            for classe in range(1, 14):
                row_idx = 20 + classe
                if row_idx < len(all_rows) and col_idx < len(all_rows[row_idx]):
                    rev_by_classe[str(classe)] = all_rows[row_idx][col_idx]
            cashback["revenda"][pci_name] = {"tipo": "por_classe", "valores": rev_by_classe}
        else:
            cashback["revenda"][pci_name] = {"tipo": "fixo", "valor": rev_val}

        if isinstance(rep_val, str) and "-" in rep_val:
            rep_by_classe = {}
            for classe in range(1, 14):
                row_idx = 34 + classe
                if row_idx < len(all_rows) and col_idx < len(all_rows[row_idx]):
                    rep_by_classe[str(classe)] = all_rows[row_idx][col_idx]
            cashback["representante"][pci_name] = {"tipo": "por_classe", "valores": rep_by_classe}
        else:
            cashback["representante"][pci_name] = {"tipo": "fixo", "valor": rep_val}

    wb.close()
    return {"pcis": list(pcis.keys()), "cashback": cashback}


CONVERTERS = {
    "Base IBGE_Area Vendedores Boxer.xlsx": ("ibge_responsaveis.json", convert_ibge_to_json),
    "BMAX CRITERIOS V2.xlsx": ("bmax_criterios.json", convert_criterios_to_json),
}


def main():
    import requests
    print("=== Sync SharePoint → JSON para Portal BMax ===")
    token = get_graph_token()
    print("MSAL: token obtido")

    DATA_DIR.mkdir(parents=True, exist_ok=True)

    xlsx_items = list_xlsx_files(token)

    if xlsx_items is not None:
        for item in xlsx_items:
            name = item["name"]
            print(f"  Encontrado: {name}")
            if name in CONVERTERS:
                json_name, converter = CONVERTERS[name]
                print(f"    Baixando {name}...")
                data = download_file_by_id(token, item["id"])
                print(f"    Convertendo → {json_name}...")
                result = converter(data)
                out_path = DATA_DIR / json_name
                out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
                print(f"    Salvo: {out_path}")
            else:
                print(f"    (sem conversor, pulando)")
    else:
        print("Usando download direto por item ID / caminho conhecido...")
        for xlsx_name, (json_name, converter) in CONVERTERS.items():
            item_id = KNOWN_FILES.get(xlsx_name)
            print(f"  Baixando {xlsx_name}...")
            if item_id:
                data = download_file_by_id(token, item_id)
            else:
                data = download_file_by_path(token, f"{SP_FOLDER}/{xlsx_name}")
            print(f"  Convertendo → {json_name}...")
            result = converter(data)
            out_path = DATA_DIR / json_name
            out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
            print(f"  Salvo: {out_path}")

    print("=== Sync concluído ===")


if __name__ == "__main__":
    main()
